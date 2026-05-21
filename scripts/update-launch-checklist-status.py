from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path("outputs/buildus_launch_ops_checklist_2026-05-18.xlsx")
OUTPUT_WORKBOOK = Path("outputs/buildus_launch_ops_checklist_2026-05-18_checked.xlsx")
SHEET_NAME = "출시전_필수"
CHECK_DATE = "2026-05-18"

UPDATES = {
    "Vercel Production alias가 최신 배포를 가리키는지 확인": {
        "status": "완료",
        "evidence": "https://buildus-care-flow.vercel.app",
        "memo": "프로덕션 배포 dpl_4qeEUKSrJaJT6hMt2h2kmPua25KT, 홈 200 확인.",
    },
    "ADMIN_PASSWORD가 운영 비밀번호로 설정되어 있는지 확인": {
        "status": "완료",
        "evidence": "https://buildus-care-flow.vercel.app/api/admin/auth",
        "memo": "운영 비밀번호로 로그인 API 200 및 admin_session 쿠키 발급 확인.",
    },
    "ADMIN_SESSION_SECRET, ADMIN_API_KEY가 프로덕션에 설정되어 있는지 확인": {
        "status": "완료",
        "evidence": "https://buildus-care-flow.vercel.app/admin/orders",
        "memo": "관리자 세션으로 주요 화면 200, admin API key로 stats/orders/data-export 200 확인.",
    },
    "Toss 운영 키/테스트 키 사용 상태 확인": {
        "status": "이슈",
        "evidence": "Vercel production env",
        "memo": "Toss 키는 존재하지만 PAYMENT_MOCK_MODE=true. 정식 결제 출시 전 false 전환 및 실제 결제 1건 재검증 필요.",
    },
    "모바일 카카오 인앱/사파리/크롬 결제창 이동 확인": {
        "status": "미확인",
        "evidence": "실기기 확인 필요",
        "memo": "API 결제 confirm은 통과. 카카오 인앱/모바일 브라우저 결제창 이동은 실제 휴대폰에서 최종 확인 필요.",
    },
    "사용자 주문 생성부터 결제완료까지 전체 플로우 테스트": {
        "status": "완료",
        "evidence": "production smoke test",
        "memo": "주문 생성, 가능한 슬롯 예약, 견적 생성/수락, 결제 confirm, 주문 상태 paid 확인. 테스트 주문 정리 완료.",
    },
    "주소 필수 입력과 견적만 확인/결제 진행 흐름 검증": {
        "status": "진행중",
        "evidence": "API smoke + UI 확인 필요",
        "memo": "API 주문 생성은 주소 포함으로 통과. 실제 UI에서 필수 표시/누락 안내/결제 진행 조건은 실기기 육안 확인 필요.",
    },
    "카카오 상담 링크와 QR 노출 방식 확인": {
        "status": "진행중",
        "evidence": "https://buildus-care-flow.vercel.app/kakao-channel-qr.png",
        "memo": "프로덕션 HTML에 카카오 링크/QR 참조 있고 QR 이미지 200 확인. 웹/모바일 반응형 노출은 최종 실기기 확인 필요.",
    },
    "주문/결제/예약 알림 큐 처리 방식 확인": {
        "status": "이슈",
        "evidence": "lib/notification-dispatcher.ts, Vercel env",
        "memo": "RESEND/SOLAPI 환경값 없음. 카카오 알림톡 provider는 미연결. 정식 출시 전 SMS/email 또는 수동 운영 절차 확정 필요.",
    },
    "주문관리 화면에서 신규/결제완료/방문예정 필터 확인": {
        "status": "진행중",
        "evidence": "https://buildus-care-flow.vercel.app/admin/orders",
        "memo": "관리자 주문관리 200 및 레이아웃 배포 확인. 실제 주문별 필터/다음 액션은 운영자 육안 확인 필요.",
    },
    "기사 배정/방문 시작/완료/검수 처리 테스트": {
        "status": "미확인",
        "evidence": "관리자/기사 실운영 시나리오 필요",
        "memo": "이번 smoke는 결제완료까지 확인. 기사 배정, 방문 시작, 완료, 검수는 별도 테스트 주문으로 검증 필요.",
    },
    "데이터 내보내기 엑셀 다운로드 확인": {
        "status": "완료",
        "evidence": "https://buildus-care-flow.vercel.app/api/admin/data-export",
        "memo": "관리자 인증 후 Excel 응답 200, content-type application/vnd.ms-excel 확인.",
    },
    "개인정보 처리방침/사업자 정보/환불 안내 노출 확인": {
        "status": "이슈",
        "evidence": "production HTML scan",
        "memo": "사업자 정보는 노출 확인. 공개 HTML에서 개인정보/환불 안내 노출은 부족하게 확인됨. 정식 출시 전 별도 고지 보강 필요.",
    },
    "관리자 페이지 비로그인 접근 차단 확인": {
        "status": "완료",
        "evidence": "https://buildus-care-flow.vercel.app/admin/orders",
        "memo": "비로그인 접근 시 /admin/login 307 리다이렉트 확인.",
    },
    "테스트 주문 데이터 정리 기준 확인": {
        "status": "완료",
        "evidence": "production smoke cleanup",
        "memo": "생성된 테스트 주문 2건 정리 완료. 테스트 데이터는 주문번호/마커 기준으로 정리 가능.",
    },
    "메인/주문완료/주문내역의 핵심 CTA 문구 확인": {
        "status": "진행중",
        "evidence": "UI 육안 확인 필요",
        "memo": "주요 페이지 200 확인. CTA 문구/시각 우선순위는 실제 PC/모바일 화면에서 최종 승인 필요.",
    },
}


def main() -> None:
    if not WORKBOOK.exists():
        raise FileNotFoundError(WORKBOOK)

    wb = load_workbook(WORKBOOK)
    ws = wb[SHEET_NAME]

    headers = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}
    required_columns = ["체크항목", "상태", "증빙링크", "메모"]
    missing = [col for col in required_columns if col not in headers]
    if missing:
        raise RuntimeError(f"Missing columns: {missing}")

    updated = 0
    for row in range(2, ws.max_row + 1):
        item = ws.cell(row=row, column=headers["체크항목"]).value
        if item not in UPDATES:
            continue
        update = UPDATES[item]
        ws.cell(row=row, column=headers["상태"]).value = update["status"]
        ws.cell(row=row, column=headers["증빙링크"]).value = update["evidence"]
        ws.cell(row=row, column=headers["메모"]).value = f"{CHECK_DATE}: {update['memo']}"
        updated += 1

    if updated != len(UPDATES):
        raise RuntimeError(f"Updated {updated} rows, expected {len(UPDATES)}")

    try:
        wb.save(WORKBOOK)
        saved_path = WORKBOOK
    except PermissionError:
        wb.save(OUTPUT_WORKBOOK)
        saved_path = OUTPUT_WORKBOOK
    print(f"updated={updated} file={saved_path.resolve()}")


if __name__ == "__main__":
    main()
