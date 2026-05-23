from __future__ import annotations

import json
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from PIL import Image, ImageOps


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT.parent / "04. 제품 리스트"
OUTPUT_JSON = ROOT / "lib" / "replacement-products.generated.json"
PUBLIC_PRODUCTS = ROOT / "public" / "products"


NS_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
PKG_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
RICH_VALUE_NS = "{http://schemas.microsoft.com/office/spreadsheetml/2017/richdata}"
MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


@dataclass(frozen=True)
class WorkbookSpec:
    path: Path
    service_code: str
    product_dir: str
    row_kind: str


SPECS = [
    WorkbookSpec(SOURCE_DIR / "세면대-사진-가격.xlsx", "basin_replace", "basins", "standard"),
    WorkbookSpec(SOURCE_DIR / "수전-사진-가격비교.xlsx", "faucet_replace", "faucets", "standard"),
    WorkbookSpec(SOURCE_DIR / "환풍기-사진-가격-2025.xlsx", "ventilator_replace", "ventilators", "ventilator"),
]


CATEGORY_META = {
    "반다리·긴다리·일체형": {
        "summary": "일반 욕실에서 가장 많이 쓰는 세면대 형태입니다. 벽 배관·하부 배관 노출을 가리고 교체 호환성이 좋습니다.",
        "decisionHint": "일반 욕실·임대·기본 교체 우선",
    },
    "탑볼·카운터형": {
        "summary": "상판 위에 올리거나 카운터와 조합하는 디자인형 세면대입니다. 하부장·상판 조건 확인이 필요합니다.",
        "decisionHint": "디자인·파우더룸·하부장 조합 우선",
    },
    "매립·언더카운터": {
        "summary": "상판에 매립하거나 아래에서 고정하는 형태입니다. 기존 상판 타공 치수와 고정 방식 확인이 중요합니다.",
        "decisionHint": "하부장·상판 타공 호환 우선",
    },
    "세면수전": {
        "summary": "세면대에 설치하는 수전입니다. 원홀·투홀·세면샤워 겸용 여부와 기존 타공 수를 확인해야 합니다.",
        "decisionHint": "욕실 세면대 교체 우선",
    },
    "주방수전": {
        "summary": "싱크대에 설치하는 원홀 주방 수전입니다. 싱크볼 타공, 하부 급수관, 샤워헤드 인출 여부를 확인합니다.",
        "decisionHint": "싱크대·주방 교체 우선",
    },
    "샤워·욕조수전": {
        "summary": "욕실 벽면에 설치하는 샤워·욕조 수전입니다. 냉온수 배관 간격과 벽면 고정 상태 확인이 필요합니다.",
        "decisionHint": "샤워 공간·욕조 수전 우선",
    },
    "기본형": {
        "summary": "단순 환기 중심의 욕실 환풍기입니다. 기존 타공 크기, 배기덕트, 전동댐퍼 필요 여부를 확인합니다.",
        "decisionHint": "임대·일반 주거·빠른 교체 우선",
    },
    "다기능": {
        "summary": "환기 외에 제습, 온풍, 스마트 기능을 더한 환풍기입니다. 전원 조건과 천장 공간 확인이 필요합니다.",
        "decisionHint": "프리미엄·다기능 욕실 우선",
    },
}


def slug(text: object) -> str:
    normalized = unicodedata.normalize("NFKC", str(text or "")).lower()
    normalized = re.sub(r"[^a-z0-9가-힣]+", "-", normalized)
    normalized = normalized.strip("-")
    return normalized or "product"


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_price(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"[^0-9]", "", str(value))
    return int(digits) if digits else None


def is_popular(note: str) -> bool:
    return "★" in note or "인기" in note or "추천" in note


def rel_targets(zip_file: zipfile.ZipFile) -> dict[str, str]:
    try:
        rv_root = ET.fromstring(zip_file.read("xl/richData/rdrichvalue.xml"))
        rich_rel_root = ET.fromstring(zip_file.read("xl/richData/richValueRel.xml"))
        rel_root = ET.fromstring(zip_file.read("xl/richData/_rels/richValueRel.xml.rels"))
    except KeyError:
        return {}

    rel_order = [rel.attrib.get(NS_REL) for rel in rich_rel_root if rel.attrib.get(NS_REL)]
    target_by_rel = {
        rel.attrib["Id"]: rel.attrib["Target"].replace("../", "xl/")
        for rel in rel_root.findall(f"{PKG_REL_NS}Relationship")
        if rel.attrib.get("Id") and rel.attrib.get("Target")
    }

    result: dict[str, str] = {}
    for index, rv in enumerate(rv_root.findall(f"{RICH_VALUE_NS}rv"), start=1):
        values = rv.findall(f"{RICH_VALUE_NS}v")
        if not values or values[0].text is None:
            continue
        rel_index = int(values[0].text)
        if 0 <= rel_index < len(rel_order):
            rel_id = rel_order[rel_index]
            if rel_id and rel_id in target_by_rel:
                result[str(index)] = target_by_rel[rel_id]
    return result


def sheet_image_vm(zip_file: zipfile.ZipFile, sheet_index: int) -> dict[int, str]:
    sheet_name = f"xl/worksheets/sheet{sheet_index}.xml"
    try:
        root = ET.fromstring(zip_file.read(sheet_name))
    except KeyError:
        return {}

    by_row: dict[int, str] = {}
    for cell in root.findall(f".//{MAIN_NS}c"):
        ref = cell.attrib.get("r", "")
        vm = cell.attrib.get("vm")
        match = re.fullmatch(r"A(\d+)", ref)
        if match and vm:
            by_row[int(match.group(1))] = vm
    return by_row


def save_image(zip_file: zipfile.ZipFile, media_path: str | None, output: Path) -> str | None:
    if not media_path:
        return None
    try:
        raw = zip_file.read(media_path)
    except KeyError:
        return None

    output.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(BytesIO(raw)) as image:
        image = ImageOps.exif_transpose(image)
        if image.mode in ("RGBA", "LA"):
            background = Image.new("RGB", image.size, (255, 255, 255))
            background.paste(image, mask=image.getchannel("A"))
            image = background
        else:
            image = image.convert("RGB")
        image.thumbnail((1100, 1100), Image.Resampling.LANCZOS)
        image.save(output, "WEBP", quality=84, method=6)
    return f"/products/{output.parent.name}/{output.name}"


def product_id(service_code: str, category: str, row_index: int, sku: str, model: str) -> str:
    return f"{service_code}:{slug(category)}:{row_index:02d}:{slug(sku or model)}"


def extract_workbook(spec: WorkbookSpec) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(spec.path, data_only=True)
    rows: list[dict[str, object]] = []
    output_dir = PUBLIC_PRODUCTS / spec.product_dir

    with zipfile.ZipFile(spec.path) as zip_file:
        media_by_vm = rel_targets(zip_file)
        sheet_vm_by_row = {
            index: sheet_image_vm(zip_file, index)
            for index in range(1, len(wb.worksheets) + 1)
        }

        for sheet_index, ws in enumerate(wb.worksheets, start=1):
            if ws.title in {"종합비교", "현장가이드", "결정 가이드"}:
                continue
            for row_index in range(4, ws.max_row + 1):
                if spec.row_kind == "ventilator":
                    category = clean(ws.cell(row_index, 2).value)
                    brand = clean(ws.cell(row_index, 3).value)
                    model = clean(ws.cell(row_index, 4).value)
                    sku = clean(ws.cell(row_index, 5).value)
                    price = clean_price(ws.cell(row_index, 6).value)
                    purpose = clean(ws.cell(row_index, 7).value)
                    note = clean(ws.cell(row_index, 8).value)
                    note = f"{purpose} · {note}" if purpose and note else purpose or note
                else:
                    category = ws.title
                    brand = clean(ws.cell(row_index, 2).value)
                    model = clean(ws.cell(row_index, 3).value)
                    sku = clean(ws.cell(row_index, 4).value)
                    price = clean_price(ws.cell(row_index, 5).value)
                    note = clean(ws.cell(row_index, 6).value)

                if not brand or not model or price is None:
                    continue

                sequence = len(rows) + 1
                pid = product_id(spec.service_code, category, sequence, sku, model)
                vm = sheet_vm_by_row.get(sheet_index, {}).get(row_index)
                media_path = media_by_vm.get(vm or "")
                image_name = f"{slug(spec.service_code)}-{sequence:02d}-{slug(sku or model)}.webp"
                image = save_image(zip_file, media_path, output_dir / image_name)
                meta = CATEGORY_META.get(category, {})

                rows.append(
                    {
                        "id": pid,
                        "serviceCode": spec.service_code,
                        "categoryId": slug(category),
                        "categoryName": category,
                        "categorySummary": meta.get("summary", ""),
                        "decisionHint": meta.get("decisionHint", ""),
                        "brand": brand,
                        "model": model,
                        "sku": sku,
                        "price": price,
                        "note": note,
                        "popular": is_popular(note),
                        "image": image,
                        "sourceWorkbook": spec.path.name,
                        "sourceSheet": ws.title,
                        "sourceRow": row_index,
                    }
                )

    return rows


def main() -> None:
    all_products: list[dict[str, object]] = []
    for spec in SPECS:
        all_products.extend(extract_workbook(spec))

    OUTPUT_JSON.write_text(json.dumps(all_products, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    by_service: dict[str, int] = {}
    for product in all_products:
        by_service[product["serviceCode"]] = by_service.get(product["serviceCode"], 0) + 1
    print(f"wrote {OUTPUT_JSON.relative_to(ROOT)}")
    print(by_service)


if __name__ == "__main__":
    main()
