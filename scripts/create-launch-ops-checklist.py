from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_DIR = Path("outputs")
OUTPUT_FILE = OUTPUT_DIR / "buildus_launch_ops_checklist_2026-05-18.xlsx"

HEADERS = ["영역", "체크항목", "출시판정", "우선순위", "담당", "상태", "점검방법", "완료기준", "증빙링크", "메모"]
STATUS_OPTIONS = ["미확인", "진행중", "완료", "보류", "이슈"]
PRIORITY_OPTIONS = ["P0", "P1", "P2"]
RELEASE_OPTIONS = ["필수", "권장"]

SHEETS = {
    "출시전_필수": [
        ["환경", "Vercel Production alias가 최신 배포를 가리키는지 확인", "필수", "P0", "운영", "미확인", "Vercel deployment/alias 확인", "buildus-care-flow.vercel.app 최신 배포 연결", "", ""],
        ["환경", "ADMIN_PASSWORD가 운영 비밀번호로 설정되어 있는지 확인", "필수", "P0", "운영", "미확인", "Vercel Environment Variables 확인", "운영자가 알고 있는 비밀번호로 로그인 성공", "", "로컬 .env와 프로덕션 값 불일치 가능성 확인 필요"],
        ["환경", "ADMIN_SESSION_SECRET, ADMIN_API_KEY가 프로덕션에 설정되어 있는지 확인", "필수", "P0", "개발", "미확인", "관리자 화면/API 접근 테스트", "관리자 화면 200, 비인증 접근 차단", "", ""],
        ["결제", "Toss 운영 키/테스트 키 사용 상태 확인", "필수", "P0", "개발", "미확인", "Vercel 환경값과 Toss 콘솔 비교", "정식 출시 시 운영 키 사용 여부 결정 완료", "", ""],
        ["결제", "모바일 카카오 인앱/사파리/크롬 결제창 이동 확인", "필수", "P0", "운영", "미확인", "실제 휴대폰에서 주문 후 결제 버튼 테스트", "모바일에서 결제창 진입 또는 안내 문구 정상", "", ""],
        ["주문", "사용자 주문 생성부터 결제완료까지 전체 플로우 테스트", "필수", "P0", "운영", "미확인", "실제 주문 1건 또는 E2E 테스트", "주문 상태가 paid까지 도달", "", ""],
        ["주문", "주소 필수 입력과 견적만 확인/결제 진행 흐름 검증", "필수", "P0", "운영", "미확인", "PC/모바일 주문 폼 테스트", "필수값 누락 시 명확한 안내, 결제 진행 시 주소 확보", "", ""],
        ["상담", "카카오 상담 링크와 QR 노출 방식 확인", "필수", "P1", "운영", "미확인", "PC/모바일 각 화면 확인", "PC는 QR, 모바일은 상담 버튼으로 노출", "", ""],
        ["알림", "주문/결제/예약 알림 큐 처리 방식 확인", "필수", "P0", "개발", "미확인", "notification_queue와 cron/process API 확인", "실제 발송 또는 운영자가 확인 가능한 큐 상태 확보", "", ""],
        ["관리자", "주문관리 화면에서 신규/결제완료/방문예정 필터 확인", "필수", "P0", "운영", "미확인", "관리자 주문관리 화면 테스트", "운영자가 다음 액션을 바로 판단 가능", "", ""],
        ["관리자", "기사 배정/방문 시작/완료/검수 처리 테스트", "필수", "P0", "운영", "미확인", "테스트 주문으로 관리자/기사 플로우 진행", "완료 및 검수 상태까지 전환 가능", "", ""],
        ["관리자", "데이터 내보내기 엑셀 다운로드 확인", "권장", "P1", "운영", "미확인", "관리자 설정 화면에서 다운로드", "마스킹/PII 포함 파일 모두 다운로드 가능", "", ""],
        ["개인정보", "개인정보 처리방침/사업자 정보/환불 안내 노출 확인", "필수", "P0", "운영", "미확인", "푸터/주문/결제 완료 화면 확인", "고객이 사업자와 취소/환불 기준을 확인 가능", "", ""],
        ["보안", "관리자 페이지 비로그인 접근 차단 확인", "필수", "P0", "개발", "미확인", "/admin/orders 직접 접근", "/admin/login으로 리다이렉트", "", ""],
        ["데이터", "테스트 주문 데이터 정리 기준 확인", "권장", "P1", "개발", "미확인", "DB에서 테스트 마커/주문번호 관리", "테스트 데이터가 운영 통계에 섞이지 않음", "", ""],
        ["콘텐츠", "메인/주문완료/주문내역의 핵심 CTA 문구 확인", "권장", "P2", "운영", "미확인", "PC/모바일 화면 육안 점검", "불필요한 문구 없이 다음 행동이 명확함", "", ""],
    ],
    "출시당일": [
        ["배포", "출시 직전 npm run typecheck, npm run build 통과 확인", "필수", "P0", "개발", "미확인", "로컬 또는 CI 빌드 로그 확인", "오류 없이 통과", "", ""],
        ["배포", "Vercel 프로덕션 배포 완료 및 alias 확인", "필수", "P0", "개발", "미확인", "vercel --prod 결과 확인", "Production alias 연결", "", ""],
        ["스모크", "홈/서비스/사진판정/주문조회 페이지 200 확인", "필수", "P0", "개발", "미확인", "프로덕션 URL 직접 호출", "주요 공개 페이지 200", "", ""],
        ["스모크", "관리자 dashboard/orders/jobs/diagnoses/settings 200 확인", "필수", "P0", "개발", "미확인", "관리자 세션으로 접근", "관리자 주요 화면 200", "", ""],
        ["스모크", "주문 1건 실제 접수 후 관리자 주문관리에서 확인", "필수", "P0", "운영", "미확인", "실제 고객처럼 주문 입력", "관리자 화면에 주문 노출", "", ""],
        ["결제", "실제 결제 또는 Toss 테스트 결제 1건 확인", "필수", "P0", "운영", "미확인", "PC/모바일 각각 테스트", "결제 완료 후 주문 상태 반영", "", ""],
        ["상담", "카카오 상담 버튼/QR 실제 연결 확인", "필수", "P1", "운영", "미확인", "휴대폰 카카오톡 앱으로 확인", "상담 채널로 정상 이동", "", ""],
        ["운영", "첫 주문 대응 담당자와 응답 시간 기준 공유", "필수", "P1", "운영", "미확인", "운영 채팅방/문서 확인", "누가 언제까지 응답할지 정해짐", "", ""],
        ["운영", "오류 발생 시 임시 접수 채널 준비", "권장", "P1", "운영", "미확인", "카카오 상담/전화/수기 접수표 준비", "웹 장애 시 주문 누락 방지", "", ""],
    ],
    "출시후_모니터링": [
        ["트래픽", "첫 24시간 방문/전환 흐름 확인", "필수", "P1", "운영", "미확인", "관리자 funnel/analytics 확인", "유입 대비 주문 전환 이상치 없음", "", ""],
        ["주문", "신규 주문이 관리자 주문관리 단계에 맞게 쌓이는지 확인", "필수", "P0", "운영", "미확인", "1일 3회 주문관리 확인", "미처리 주문 없음", "", ""],
        ["결제", "결제 실패/중복 결제/취소 요청 모니터링", "필수", "P0", "운영", "미확인", "payments/cancellations 확인", "문제 주문 당일 처리", "", ""],
        ["예약", "슬롯 마감/차단/배정 가능 인원 확인", "필수", "P1", "운영", "미확인", "관리자 슬롯/기사 화면 확인", "예약 가능 수량과 기사 수가 맞음", "", ""],
        ["상담", "카카오 상담 문의 유형 기록", "권장", "P2", "운영", "미확인", "문의 사유 태깅", "자주 묻는 질문/폼 개선 항목 도출", "", ""],
        ["성능", "모바일 주요 화면 로딩과 깨짐 확인", "권장", "P1", "개발", "미확인", "실기기/브라우저 개발자 도구 확인", "가로 스크롤/텍스트 겹침 없음", "", ""],
        ["데이터", "일일 데이터 내보내기 백업 확인", "권장", "P1", "운영", "미확인", "관리자 설정에서 다운로드", "주문/결제/예약 데이터 백업 가능", "", ""],
    ],
    "장애대응": [
        ["결제 장애", "결제창 이동 실패 시 고객 안내 문구/상담 전환", "필수", "P0", "운영", "미확인", "모바일 결제 실패 시나리오 테스트", "상담으로 결제 안내 가능", "", ""],
        ["주문 장애", "주문 제출 실패 시 임시 접수 경로 운영", "필수", "P0", "운영", "미확인", "카카오 상담/수기 접수표 확인", "고객 정보와 서비스 요청을 누락 없이 수집", "", ""],
        ["관리자 장애", "관리자 로그인 불가 시 환경값/세션 시크릿 확인 절차", "필수", "P0", "개발", "미확인", "Vercel env와 배포 로그 확인", "30분 내 복구 또는 우회 처리", "", ""],
        ["DB 장애", "Supabase 연결 오류 시 확인 순서", "필수", "P0", "개발", "미확인", "Supabase 상태/키/네트워크 확인", "원인과 복구 예정 시간 공유", "", ""],
        ["알림 장애", "알림 발송 실패 시 관리자 수동 확인 프로세스", "필수", "P1", "운영", "미확인", "주문관리 미처리 필터 확인", "알림 없이도 주문 대응 가능", "", ""],
        ["환불/취소", "취소 요청 접수 후 승인/반려/환불 기록 절차", "필수", "P1", "운영", "미확인", "관리자 취소 처리 화면 확인", "고객 안내와 내부 기록 일치", "", ""],
    ],
    "계정_환경값": [
        ["Vercel", "프로덕션 프로젝트 접근 권한", "필수", "P0", "개발", "미확인", "Vercel 팀/계정 확인", "최소 2명 이상 접근 가능", "", ""],
        ["Supabase", "DB/Storage 접근 권한", "필수", "P0", "개발", "미확인", "Supabase 프로젝트 권한 확인", "운영 담당 접근 가능", "", ""],
        ["Toss", "결제 콘솔/키 관리 권한", "필수", "P0", "운영", "미확인", "Toss 콘솔 접근 확인", "운영/정산 담당 접근 가능", "", ""],
        ["Kakao", "상담 채널 관리자 권한", "필수", "P1", "운영", "미확인", "카카오 채널 관리자 확인", "문의 응대 담당 접근 가능", "", ""],
        ["도메인", "대표 도메인 연결 계획", "권장", "P2", "운영", "미확인", "도메인/DNS 확인", "정식 도메인 또는 Vercel 주소 사용 방침 결정", "", ""],
        ["문서", "운영 매뉴얼/데이터 수집 매뉴얼 보관 위치", "권장", "P2", "운영", "미확인", "문서 링크 확인", "신규 담당자가 바로 확인 가능", "", ""],
    ],
}


def style_sheet(ws, rows_count: int) -> None:
    widths = {
        "A": 14,
        "B": 38,
        "C": 11,
        "D": 10,
        "E": 12,
        "F": 12,
        "G": 32,
        "H": 32,
        "I": 24,
        "J": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{rows_count}"

    header_fill = PatternFill("solid", fgColor="1C1B19")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="E8E5E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=rows_count, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        row[2].alignment = Alignment(horizontal="center", vertical="top")
        row[3].alignment = Alignment(horizontal="center", vertical="top")
        row[5].alignment = Alignment(horizontal="center", vertical="top")

    status_dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=False)
    priority_dv = DataValidation(type="list", formula1=f'"{",".join(PRIORITY_OPTIONS)}"', allow_blank=False)
    release_dv = DataValidation(type="list", formula1=f'"{",".join(RELEASE_OPTIONS)}"', allow_blank=False)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(priority_dv)
    ws.add_data_validation(release_dv)
    status_dv.add(f"F2:F{rows_count}")
    priority_dv.add(f"D2:D{rows_count}")
    release_dv.add(f"C2:C{rows_count}")

    ws.conditional_formatting.add(f"F2:F{rows_count}", FormulaRule(formula=['$F2="완료"'], fill=PatternFill("solid", fgColor="DCFCE7")))
    ws.conditional_formatting.add(f"F2:F{rows_count}", FormulaRule(formula=['$F2="이슈"'], fill=PatternFill("solid", fgColor="FEE2E2")))
    ws.conditional_formatting.add(f"F2:F{rows_count}", FormulaRule(formula=['$F2="진행중"'], fill=PatternFill("solid", fgColor="DBEAFE")))
    ws.conditional_formatting.add(f"F2:F{rows_count}", FormulaRule(formula=['$F2="보류"'], fill=PatternFill("solid", fgColor="FFEDD5")))
    ws.conditional_formatting.add(f"C2:C{rows_count}", FormulaRule(formula=['$C2="필수"'], fill=PatternFill("solid", fgColor="F8FBFF")))
    ws.conditional_formatting.add(f"D2:D{rows_count}", FormulaRule(formula=['$D2="P0"'], fill=PatternFill("solid", fgColor="FFF1F2")))


def build_workbook() -> Workbook:
    wb = Workbook()
    summary = wb.active
    summary.title = "요약"

    summary["A1"] = "Buildus Care 운영 출시 체크리스트"
    summary["A2"] = "상태는 각 시트의 드롭다운으로 관리합니다. 필수/P0 항목은 정식 출시 전 반드시 완료 또는 명시적 보류 사유를 남겨야 합니다."
    summary["A1"].font = Font(size=18, bold=True, color="1C1B19")
    summary["A2"].font = Font(size=10, color="5C5A54")
    summary.merge_cells("A1:H1")
    summary.merge_cells("A2:H2")

    summary_headers = ["시트", "총 항목", "완료", "진행중", "이슈", "보류", "완료율", "출시판정"]
    for col, value in enumerate(summary_headers, start=1):
        cell = summary.cell(row=4, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor="1C1B19")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for index, sheet_name in enumerate(SHEETS.keys(), start=5):
        summary.cell(index, 1, sheet_name)
        summary.cell(index, 2, f'=COUNTA(\'{sheet_name}\'!B2:B200)')
        summary.cell(index, 3, f'=COUNTIF(\'{sheet_name}\'!F2:F200,"완료")')
        summary.cell(index, 4, f'=COUNTIF(\'{sheet_name}\'!F2:F200,"진행중")')
        summary.cell(index, 5, f'=COUNTIF(\'{sheet_name}\'!F2:F200,"이슈")')
        summary.cell(index, 6, f'=COUNTIF(\'{sheet_name}\'!F2:F200,"보류")')
        summary.cell(index, 7, f'=IF(B{index}=0,0,C{index}/B{index})')
        summary.cell(index, 8, f'=IF(E{index}>0,"이슈확인",IF(G{index}=1,"완료","점검중"))')

    total_row = 5 + len(SHEETS)
    summary.cell(total_row, 1, "전체")
    for col in range(2, 7):
        summary.cell(total_row, col, f"=SUM({chr(64+col)}5:{chr(64+col)}{total_row-1})")
    summary.cell(total_row, 7, f"=IF(B{total_row}=0,0,C{total_row}/B{total_row})")
    summary.cell(total_row, 8, f'=IF(E{total_row}>0,"이슈확인",IF(G{total_row}=1,"출시가능","점검필요"))')

    summary["A12"] = "출시 Go/No-Go 기준"
    summary["A13"] = "1. 필수/P0 항목은 완료가 원칙입니다."
    summary["A14"] = "2. 보류 항목은 보류 사유와 임시 운영 절차가 있어야 합니다."
    summary["A15"] = "3. 이슈 항목이 1개 이상이면 정식 출시 전 원인/담당/복구 시간을 확정합니다."
    summary["A16"] = "4. 결제/알림/관리자 로그인은 실제 운영 환경에서 재확인합니다."
    summary["A12"].font = Font(bold=True, color="01696F")

    for col, width in {"A": 22, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 14}.items():
        summary.column_dimensions[col].width = width
    for row in summary.iter_rows(min_row=4, max_row=total_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = Border(left=Side(style="thin", color="E8E5E0"), right=Side(style="thin", color="E8E5E0"), top=Side(style="thin", color="E8E5E0"), bottom=Side(style="thin", color="E8E5E0"))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(5, total_row + 1):
        summary.cell(row, 7).number_format = "0%"
    summary.cell(total_row, 1).font = Font(bold=True)
    for col in range(1, 9):
        summary.cell(total_row, col).fill = PatternFill("solid", fgColor="F0FDF4")
        summary.cell(total_row, col).font = Font(bold=True)

    chart = BarChart()
    chart.title = "시트별 완료 현황"
    chart.y_axis.title = "항목 수"
    chart.x_axis.title = "체크리스트"
    data = Reference(summary, min_col=3, max_col=5, min_row=4, max_row=total_row - 1)
    cats = Reference(summary, min_col=1, min_row=5, max_row=total_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 16
    summary.add_chart(chart, "J4")

    for sheet_name, rows in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
        for row in rows:
            ws.append(row)
        style_sheet(ws, len(rows) + 1)

    wb.calculation.fullCalcOnLoad = True
    return wb


def verify(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    required = ["요약", *SHEETS.keys()]
    missing = [name for name in required if name not in wb.sheetnames]
    if missing:
        raise RuntimeError(f"Missing sheets: {missing}")
    for sheet_name in SHEETS:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            raise RuntimeError(f"{sheet_name} has no checklist rows")
        if [ws.cell(1, col).value for col in range(1, len(HEADERS) + 1)] != HEADERS:
            raise RuntimeError(f"{sheet_name} header mismatch")
    formula_cells = [cell.value for row in wb["요약"].iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    if not formula_cells:
        raise RuntimeError("Summary formulas were not written")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUTPUT_FILE)
    verify(OUTPUT_FILE)
    print(OUTPUT_FILE.resolve())


if __name__ == "__main__":
    main()
