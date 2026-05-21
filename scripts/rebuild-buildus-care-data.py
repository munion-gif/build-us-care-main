from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "buildus_care_data.xlsx"

GREEN = "16735F"
LIGHT_GREEN = "E7F1EB"
LIGHT_GRAY = "F4F6F3"
YELLOW = "FFF3C4"
RED = "FBE3E3"
BORDER = Border(bottom=Side(style="thin", color="D9DED5"))


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[object]] | None = None, widths: dict[int, int] | None = None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows or []:
        ws.append(row)
    style_header(ws)
    apply_table(ws, title)
    apply_widths(ws, widths)
    ws.freeze_panes = "A2"
    return ws


def style_header(ws):
    header_fill = PatternFill("solid", fgColor=GREEN)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    ws.sheet_view.showGridLines = False


def apply_table(ws, name: str):
    if ws.max_column < 1 or ws.max_row < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 2)}"
    table_name = "".join(ch if ch.isalnum() else "_" for ch in f"tbl_{name}")[:250]
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def apply_widths(ws, widths: dict[int, int] | None = None):
    default_widths = widths or {}
    for idx in range(1, ws.max_column + 1):
      width = default_widths.get(idx)
      if width is None:
          sample = [str(ws.cell(row=row, column=idx).value or "") for row in range(1, min(ws.max_row, 20) + 1)]
          width = min(max(max((len(item) for item in sample), default=8) + 2, 12), 34)
      ws.column_dimensions[get_column_letter(idx)].width = width


def add_validation(ws, column_letter: str, values: list[str], start_row: int = 2, end_row: int = 500):
    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{column_letter}{start_row}:{column_letter}{end_row}")


def add_readme(wb: Workbook):
    ws = wb.active
    ws.title = "README"
    rows = [
        ["buildus care 데이터 수집 워크북", "2026-05-18 기준"],
        ["목적", "광고/퍼널만 보던 파일을 고객-주거-주문-견적-결제-예약-현장-사진-검수-A/S까지 이어지는 운영 데이터셋으로 재정리"],
        ["수집 원칙", "PII 분리, 단계별 최소 수집, 사진 표준 컷, 현장 기사 부담 최소화, 이벤트 로그와 운영 테이블 분리"],
        ["P0 우선", "customers_private, homes, orders, quotes, payments, reservations, jobs, media, events, sessions"],
        ["P1 확장", "inspections, feedbacks, warranty_cases, notification_queue, materials, technicians"],
        ["사용 방식", "실제 서비스 DB와 동일한 키를 유지하고, 개인정보 시트는 접근 제한/암호화 대상"],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(size=16, bold=True, color=GREEN)
    ws["B1"].font = Font(size=12, bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 100
    ws.sheet_view.showGridLines = False


def main():
    wb = Workbook()
    add_readme(wb)

    add_sheet(
        wb,
        "collection_map",
        ["stage", "priority", "domain", "table", "field_group", "collection_point", "current_app_status", "note"],
        [
            ["Stage 0", "P0", "고객/PII", "customers_private", "이름/전화/카톡동의", "사진판정/견적/주문", "부분 구현", "넓은 관리자 화면은 마스킹, 상세/처리 화면만 원문 접근"],
            ["Stage 0", "P0", "주거 정보", "homes", "주소/동/아파트/주거형태/평수", "결제 전 주문 단계", "구현", "견적만 볼 때는 최소화, 결제/예약 진행 시 필수"],
            ["Stage 0", "P0", "주문", "orders", "서비스/사유/긴급도/상태/채널", "견적 생성", "구현", "service_code는 underscore 규칙 사용"],
            ["Stage 0", "P0", "견적/결제", "quotes/payments", "견적금액/옵션/결제상태", "견적/토스결제", "구현", "카드 정보 저장 금지"],
            ["Stage 0", "P0", "예약/현장", "reservations/jobs", "방문일/시간/기사/상태", "결제 후 예약/관리자 배정", "구현", "기사 PII는 방문 전 필요한 범위만 노출"],
            ["Stage 1", "P1", "사진", "media", "표준 컷/각도/태그", "사진판정/현장사진", "보강 완료", "고객: wide/close/context, 현장: before/during/after/material/issue"],
            ["Stage 1", "P1", "품질", "inspections/feedbacks", "검수/NPS/리뷰", "작업 완료 후", "부분 구현", "검수 실패 사유와 사진 연결 필요"],
            ["Stage 1", "P1", "A/S", "warranty_cases", "증상/책임/처리상태", "주문내역/A/S 접수", "부분 구현", "만료 알림과 연결"],
            ["Stage 2", "P2", "마케팅", "events/sessions", "UTM/이벤트/디바이스", "전체 플로우", "구현", "운영 테이블과 별도 관리"],
        ],
        {1: 12, 2: 10, 3: 14, 4: 22, 5: 26, 6: 24, 7: 16, 8: 50},
    )

    add_sheet(
        wb,
        "data_dictionary",
        ["table", "field", "type", "required", "pii", "priority", "source", "description"],
        [
            ["customers_private", "customer_id", "uuid/text", "Y", "N", "P0", "server", "고객 내부 식별자"],
            ["customers_private", "name", "text", "Y", "Y", "P0", "user", "고객명"],
            ["customers_private", "phone", "text", "Y", "Y", "P0", "user", "010 휴대폰 번호"],
            ["customers_private", "kakao_notice_consent", "boolean", "Y", "Y", "P0", "user", "카카오 알림/상담 동의"],
            ["homes", "address_full", "text", "Y", "Y", "P0", "user/address api", "결제/예약 진행 시 방문 주소"],
            ["homes", "housing_type", "enum", "Y", "N", "P0", "user", "자가/전세/월세/기타"],
            ["homes", "building_type", "enum", "Y", "N", "P0", "user", "아파트/빌라/단독/오피스텔"],
            ["orders", "reason", "text", "Y", "N", "P0", "user", "교체/수리 요청 사유"],
            ["orders", "urgency", "enum", "Y", "N", "P0", "user", "normal/urgent/scheduled"],
            ["media", "angle", "enum", "Y", "N", "P1", "upload ui", "wide/close/context/before/during/after/material/issue"],
            ["jobs", "materials_used", "json", "N", "N", "P1", "technician", "사용 자재 SKU와 수량"],
            ["inspections", "checklist_results", "json", "Y", "N", "P1", "admin", "검수 체크리스트 결과"],
            ["events", "event_type", "enum", "Y", "N", "P2", "client/server", "퍼널/행동 이벤트"],
        ],
        {1: 22, 2: 26, 3: 16, 4: 10, 5: 8, 6: 10, 7: 18, 8: 54},
    )

    add_sheet(
        wb,
        "customers_private",
        ["customer_id", "created_at", "name", "phone", "kakao_notice_consent", "acquisition_source", "first_contact_at", "retention_until", "access_level"],
        [["CUST-0001", "2026-05-12 10:03:45", "김OO", "01012345678", True, "instagram", "2026-05-12 10:11:45", "2029-05-12", "restricted"]],
        {1: 16, 2: 20, 3: 14, 4: 16, 5: 18, 6: 18, 7: 20, 8: 16, 9: 14},
    )

    add_sheet(
        wb,
        "homes",
        ["home_id", "customer_id", "address_full", "address_dong", "address_apt", "postal_code", "housing_type", "building_type", "size_pyung", "year_built", "floor", "has_kids", "has_elderly"],
        [["HOME-0001", "CUST-0001", "경기도 수원시 영통구 예시로 101", "영통동", "예시아파트", "16600", "자가", "아파트", 32, 2012, 12, False, False]],
    )

    add_sheet(
        wb,
        "orders",
        ["order_id", "order_number", "customer_id", "home_id", "created_at", "channel", "source", "campaign", "service_code", "skus", "reason", "urgency", "self_diagnosis", "status", "total_amount", "access_token_issued"],
        [["ORD-0001", "BU-20260512-0001", "CUST-0001", "HOME-0001", "2026-05-12 10:03:45", "web", "instagram", "suwon_toilet_fixed_price", "toilet_replace", "toilet_replace:standard", "누수와 흔들림", "normal", "사진 기준 교체추천", "paid", 325000, True]],
        {1: 14, 2: 20, 11: 28, 13: 26, 16: 18},
    )

    add_sheet(
        wb,
        "quotes",
        ["quote_id", "order_id", "created_at", "service_code", "base_amount", "materials_amount", "addon_amount", "discount_amount", "final_amount", "accepted_at", "quote_status"],
        [["QUOTE-0001", "ORD-0001", "2026-05-12 10:05:20", "toilet_replace", 250000, 60000, 15000, 0, 325000, "2026-05-12 10:08:00", "accepted"]],
    )

    add_sheet(
        wb,
        "payments",
        ["payment_id", "order_id", "quote_id", "provider", "payment_key", "method", "amount", "status", "paid_at", "refund_amount", "provider_status"],
        [["PAY-0001", "ORD-0001", "QUOTE-0001", "toss", "masked_test_key", "CARD", 325000, "done", "2026-05-12 10:09:10", 0, "DONE"]],
    )

    add_sheet(
        wb,
        "reservations",
        ["reservation_id", "order_id", "reserved_date", "time_slot", "status", "created_at", "rescheduled_count", "memo"],
        [["RSV-0001", "ORD-0001", "2026-05-14", "morning", "confirmed", "2026-05-12 10:09:50", 0, "엘리베이터 사용 가능"]],
    )

    add_sheet(
        wb,
        "jobs",
        ["job_id", "order_id", "technician_id", "scheduled_at", "started_at", "ended_at", "status", "expected_minutes", "actual_minutes", "materials_used", "extra_materials", "issues", "completion_notes"],
        [["JOB-0001", "ORD-0001", "TECH-0001", "2026-05-14 09:00:00", "2026-05-14 09:08:00", "2026-05-14 10:12:00", "done", 60, 64, '[{\"sku\":\"angle_valve\",\"qty\":1}]', "", "", "누수 확인 완료"]],
        {10: 30, 11: 24, 12: 24, 13: 30},
    )

    add_sheet(
        wb,
        "media",
        ["media_id", "order_id", "job_id", "created_at", "type", "angle", "file_path", "url", "tags", "sort_order", "ai_detected", "exif_removed"],
        [
            ["MED-0001", "ORD-0001", "", "2026-05-12 10:04:10", "inquiry", "wide", "orders/ORD-0001/inquiry/wide.jpg", "", "customer,wide", 1, "", True],
            ["MED-0002", "ORD-0001", "", "2026-05-12 10:04:20", "inquiry", "close", "orders/ORD-0001/inquiry/close.jpg", "", "customer,close", 2, "", True],
            ["MED-0003", "ORD-0001", "", "2026-05-12 10:04:30", "inquiry", "context", "orders/ORD-0001/inquiry/context.jpg", "", "customer,context", 3, "", True],
        ],
        {7: 36, 8: 22, 9: 18, 11: 24},
    )

    add_sheet(
        wb,
        "inspections",
        ["inspection_id", "job_id", "order_id", "created_at", "passed", "checklist_results", "inspector_note"],
        [["INS-0001", "JOB-0001", "ORD-0001", "2026-05-14 11:00:00", True, '[{\"item\":\"시공 전후 사진 일치\",\"ok\":true}]', "이상 없음"]],
        {6: 46, 7: 30},
    )

    add_sheet(
        wb,
        "feedbacks",
        ["feedback_id", "order_id", "customer_id", "created_at", "nps", "rating", "comment", "public_review_agreed"],
        [["FDB-0001", "ORD-0001", "CUST-0001", "2026-05-14 18:00:00", 9, 5, "시간 맞춰 깔끔하게 완료", True]],
    )

    add_sheet(
        wb,
        "warranty_cases",
        ["warranty_id", "order_id", "job_id", "created_at", "issue_type", "description", "responsibility", "status", "resolved_at", "resolution_note"],
        [["WAR-0001", "ORD-0001", "JOB-0001", "", "leak", "", "", "none", "", ""]],
    )

    add_sheet(
        wb,
        "notification_queue",
        ["notification_id", "order_id", "customer_id", "channel", "template_key", "scheduled_at", "sent_at", "status", "error_message"],
        [["NOTI-0001", "ORD-0001", "CUST-0001", "kakao", "payment_complete", "2026-05-12 10:09:20", "", "queued", ""]],
    )

    add_sheet(
        wb,
        "events",
        ["event_id", "occurred_at", "session_id", "customer_id", "order_id", "event_type", "source", "campaign", "landing_path", "device_type", "service_code", "properties"],
        [["EVT-0001", "2026-05-12 10:03:21", "SES-0001", "", "", "landing_view", "instagram", "suwon_toilet_fixed_price", "/quote/toilet_replace", "mobile", "toilet_replace", "{}"]],
        {12: 36},
    )

    add_sheet(
        wb,
        "sessions",
        ["session_id", "first_event_time", "last_event_time", "source", "campaign", "landing_path", "device_type", "region_hint", "order_id", "utm_source", "utm_campaign", "referrer"],
        [["SES-0001", "2026-05-12 10:03:21", "2026-05-12 10:09:10", "instagram", "suwon_toilet_fixed_price", "/quote/toilet_replace", "mobile", "suwon_yeongtong", "ORD-0001", "instagram", "suwon_toilet_fixed_price", ""]],
    )

    add_sheet(
        wb,
        "materials",
        ["sku", "name", "category", "grade", "retail_price", "cost_price", "is_active", "default_service_code"],
        [["angle_valve", "앵글밸브", "설비자재", "standard", 15000, 8000, True, "toilet_replace"]],
    )

    add_sheet(
        wb,
        "technicians",
        ["technician_id", "name_masked", "phone_masked", "region", "type", "grade", "skills", "avg_nps", "pass_rate", "is_active"],
        [["TECH-0001", "김**", "010-1234-****", "수원·용인", "contractor", "silver", "toilet_replace,faucet_replace", 9.1, 98, True]],
    )

    add_sheet(wb, "dim_services", ["code", "name", "standardizable", "description"], [
        ["toilet_replace", "양변기 교체", True, "정찰가형 교체 서비스"],
        ["faucet_replace", "수전 교체", True, "싱크대/세면대 수전 교체"],
        ["drain_clog", "하수구 막힘", False, "상담형 현장 확인 서비스"],
    ])
    add_sheet(wb, "dim_channels", ["code", "name", "description"], [
        ["instagram", "인스타그램 광고", "유료 소셜 유입"],
        ["kakao", "카카오 상담", "카카오 채널 1:1 상담"],
        ["web", "웹 주문", "홈페이지 직접 주문"],
        ["phone", "전화", "전화 상담/접수"],
        ["offline", "오프라인", "소개/전단/현장"],
    ])
    add_sheet(wb, "dim_campaigns", ["code", "name", "description"], [
        ["suwon_toilet_fixed_price", "수원 양변기 교체 정찰가", "인스타/웹 정찰가 캠페인"],
        ["photo_diagnosis", "사진 판정", "사진 3장 선판정 캠페인"],
    ])
    add_sheet(wb, "dim_regions", ["code", "name", "description"], [
        ["suwon_yeongtong", "수원 영통구", "영통·광교·망포"],
        ["yongin_suji", "용인 수지구", "상현·성복·풍덕천"],
    ])
    add_sheet(wb, "dim_photo_angles", ["angle", "label", "owner", "description"], [
        ["wide", "전체 컷", "customer", "설치 위치와 주변 전체"],
        ["close", "문제 부위", "customer", "파손/누수/오작동 부위 클로즈업"],
        ["context", "주변·규격", "customer", "배관/벽/바닥/규격 정보"],
        ["before", "시공 전", "technician", "작업 전 상태"],
        ["during", "작업 중", "technician", "철거/설치 중 핵심 장면"],
        ["after", "완료 후", "technician", "완료 전체/마감"],
        ["material", "자재", "technician", "사용 자재/영수증"],
        ["issue", "이슈", "technician", "추가 비용/파손/특이사항"],
    ])

    validations = {
        "customers_private": {"E": ["TRUE", "FALSE"], "I": ["restricted", "admin_only"]},
        "homes": {"G": ["자가", "전세", "월세", "기타"], "H": ["아파트", "빌라", "단독주택", "오피스텔"]},
        "orders": {"F": ["web", "kakao", "phone", "store", "instagram", "offline"], "L": ["normal", "urgent", "scheduled"], "N": ["inquiry", "submitted", "quoted", "payment_pending", "paid", "scheduled", "in_progress", "done", "completed", "cancel_requested", "issue", "warranty"]},
        "reservations": {"D": ["morning", "afternoon", "all_day"], "E": ["pending", "confirmed", "cancelled"]},
        "payments": {"F": ["CARD", "TRANSFER", "CASH"], "H": ["pending", "done", "failed", "cancelled", "refunded"]},
        "media": {"E": ["inquiry", "before", "during", "after", "material", "issue"], "F": ["wide", "close", "context", "before", "during", "after", "material", "issue"]},
        "notification_queue": {"D": ["kakao", "sms", "email"], "H": ["queued", "processing", "sent", "failed", "cancelled"]},
    }
    for sheet, cols in validations.items():
        ws = wb[sheet]
        for col, values in cols.items():
            add_validation(ws, col, values)

    dashboard = wb.create_sheet("dashboard")
    dashboard.append(["지표", "값", "설명"])
    dashboard.append(["주문 수", '=COUNTA(orders!A2:A500)', "orders 기준"])
    dashboard.append(["결제 완료", '=COUNTIF(payments!H2:H500,"done")', "payments.status=done"])
    dashboard.append(["사진 수집", '=COUNTA(media!A2:A500)', "media 기준"])
    dashboard.append(["검수 통과", '=COUNTIF(inspections!E2:E500,TRUE)', "inspections.passed"])
    dashboard.append(["알림 대기", '=COUNTIF(notification_queue!H2:H500,"queued")', "notification_queue.status=queued"])
    style_header(dashboard)
    apply_widths(dashboard, {1: 18, 2: 16, 3: 38})
    dashboard.sheet_view.showGridLines = False

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(OUTPUT)


if __name__ == "__main__":
    main()
